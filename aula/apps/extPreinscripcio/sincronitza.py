# This Python file uses the following encoding: utf-8

from aula.apps.extPreinscripcio.models import Preinscripcio

from aula.apps.missatgeria.missatges_a_usuaris import tipusMissatge, IMPORTACIO_PREINSCRIPCIO_FINALITZADA
from aula.apps.missatgeria.models import Missatge

from openpyxl import load_workbook

from django.contrib.auth.models import Group


def assignaDades(preinscripcio, index, value, colnames, col_indexs):
    if index in col_indexs and value:
        field=colnames[col_indexs[index]]
        if isinstance(field, dict):
            if 'append' in field and field['append']:
                if field['field'] not in preinscripcio:
                    preinscripcio[field['field']]=value
                else:
                    preinscripcio[field['field']]=str(preinscripcio[field['field']])+" "+str(value)
            else:
                if 'field' in field:
                    preinscripcio[field['field']]=value
        else:
            preinscripcio[field]=value
    return preinscripcio

def sincronitza(f, user = None):
    
    errors = []
    warnings= []
    infos= []
    
    try:
        # Carregar full de càlcul
        wb2 = load_workbook(f)
        if len(wb2.worksheets)!=1:
            # Si té més d'una pestanya --> error
            errors.append('Fitxer incorrecte sheets')
            return {'errors': errors, 'warnings': [], 'infos': []}
    except:
        errors.append('Fitxer incorrecte')
        return {'errors': errors, 'warnings': [], 'infos': []}
    
    info_nAlumnesLlegits=0
    
    # Carregar full de càlcul
    full = wb2.active
    max_row = full.max_row

    # columnes que s'importaran,  camp excel : camp base de dades 
    colnames = {
        'Estat sol·licitud':{},
        'Nom': 'nom',
        'Primer cognom': 'cognoms',
        'Segon cognom': {'field': 'cognoms', 'append' : True},
        'Identificació RALC': 'ralc',
        'Codi ensenyament P1': 'codiestudis',
        'Nom ensenyament P1': 'nomestudis',
        'Codi modalitat': 'codimodalitat',
        'Modalitat': 'nommodalitat',
        'Curs P1': 'curs',
        'Règim P1': 'regim',
        'Torn P1': 'torn',
        'DNI': 'identificador',
        'NIE': {'field': 'identificador'},
        'PASS': {'field': 'identificador'},
        'TIS': 'tis',
        'Data naixement': 'naixement',
        'Sexe': 'sexe',
        'Nacionalitat': 'nacionalitat',
        'País naixement': 'paisnaixement',
        'Tipus via': 'adreça',
        'Nom via': {'field': 'adreça', 'append' : True},
        'Número via': {'field': 'adreça', 'append' : True},
        'Altres dades': {'field': 'adreça', 'append' : True},
        'Província residència': 'provincia',
        'Municipi residència': 'municipi',
        'Localitat residència': 'localitat',
        'CP': 'cp',
        'País residència': 'paisresidencia',
        'Telèfon': 'telefon',
        'Correu electrònic': 'correu',
        'Tipus doc. tutor 1': 'tdoctut1',
        'Núm. doc. tutor 1': 'doctut1',
        'Nom tutor 1': 'nomtut1',
        'Primer cognom tutor 1': 'cognomstut1',
        'Segon cognom tutor 1': {'field': 'cognomstut1', 'append' : True},
        'Tipus doc. tutor 2': 'tdoctut2',
        'Núm. doc. tutor 2': 'doctut2',
        'Nom tutor 2': 'nomtut2',
        'Primer cognom tutor 2': 'cognomstut2',
        'Segon cognom tutor 2': {'field': 'cognomstut2', 'append' : True},
        'Codi centre proc.': 'codicentreprocedencia',
        'Nom centre proc.': 'centreprocedencia',
        'Codi ensenyament proc.': 'codiestudisprocedencia',
        'Nom ensenyament proc.': 'estudisprocedencia',
        'Curs proc.': 'cursestudisprocedencia',
        }
    
    rows = list(wb2.active.rows)
    col_indexs = {n: cell.value for n, cell in enumerate(rows[0])
                   if cell.value in colnames}
    
    for row in rows[1:max_row]:
       
        preinscripcio={}
        estat=None
        for index, cell in enumerate(row):
            if bool(cell) and bool(cell.value) and isinstance(cell.value, str):
                cell.value=cell.value.strip()
            preinscripcio=assignaDades(preinscripcio, index, cell.value, colnames, col_indexs)
            if index in col_indexs and col_indexs[index]=='Estat sol·licitud':
                estat=cell.value

        if 'cp' in preinscripcio and preinscripcio['cp'].startswith("="): preinscripcio['cp']=preinscripcio['cp'][2:7]

        if estat=='Validada':
            try:
                p=Preinscripcio(**preinscripcio)
                query=Preinscripcio.objects.filter(ralc=p.ralc)
                if query:
                    p.pk=query[0].pk
                p.save()
                info_nAlumnesLlegits += 1
            except Exception as e:
                errors.append(str(e)+": "+preinscripcio)
                

    infos.append(u'{0} alumnes llegits'.format(info_nAlumnesLlegits) )

    missatge = IMPORTACIO_PREINSCRIPCIO_FINALITZADA
    tipus_de_missatge = tipusMissatge(missatge)
    msg = Missatge(
                remitent= user,
                text_missatge = missatge,
                tipus_de_missatge = tipus_de_missatge)
    msg.afegeix_errors( errors )
    msg.afegeix_warnings(warnings)
    msg.afegeix_infos(infos)
    importancia = 'VI' if len( errors )> 0 else 'IN'
    grupDireccio =  Group.objects.get( name = 'direcció' )
    msg.envia_a_grup( grupDireccio , importancia=importancia)

    return { 'errors': errors, 'warnings': warnings, 'infos': infos }
