# This Python file uses the following encoding: utf-8

#--
from aula.apps.alumnes.models import Alumne, Grup, Nivell, DadesAddicionalsAlumne
from aula.apps.alumnes.tools_aruco import markers_disponibles, set_aruco_marker
from aula.apps.missatgeria.missatges_a_usuaris import ALUMNES_DONATS_DE_BAIXA, tipusMissatge, ALUMNES_CANVIATS_DE_GRUP, \
    ALUMNES_DONATS_DALTA, IMPORTACIO_ESFERA_FINALITZADA, IMPORTACIO_DADES_ADDICIONALS_FINALITZADA
from aula.apps.presencia.models import ControlAssistencia
from aula.apps.missatgeria.models import Missatge
from aula.apps.usuaris.models import Professor
from aula.apps.relacioFamilies.tools import creaResponsables
from aula.apps.extEsfera.models import ParametreEsfera
from openpyxl import load_workbook

from django.db.models import Q

from datetime import datetime, date
from django.contrib.auth.models import Group

import time
from aula.apps.extEsfera.models import Grup2Aula
from aula.settings import CUSTOM_DADES_ADDICIONALS_ALUMNE

from aula.utils.tools import unicode
from aula.apps.extSaga.sincronitzaSaga import posarDada

def sincronitza(f, user = None):

    errors = []
    markers_per_nivell = markers_disponibles()

    try:
        # Carregar full de càlcul
        wb2 = load_workbook(f)
        if len(wb2.worksheets)!=1:
            # Si té més d'una pestanya --> error
            errors.append('Fitxer incorrecte')
            return {'errors': errors, 'warnings': [], 'infos': []}
        msgs = comprovar_grups( f )
        if msgs["errors"]:
            return msgs    
    except:
        errors.append('Fitxer incorrecte')
        return {'errors': errors, 'warnings': [], 'infos': []}

    #Exclou els alumnes AMB esborrat i amb estat MAN (creats manualment)
    Alumne.objects.exclude( estat_sincronitzacio__exact = 'DEL' ).exclude( estat_sincronitzacio__exact = 'MAN') \
        .update( estat_sincronitzacio = 'PRC')

    errors_nAlumnesSenseGrup=0
    info_nAlumnesLlegits=0
    info_nAlumnesInsertats=0
    info_nAlumnesEsborrats=0
    info_nAlumnesCanviasDeGrup=0
    info_nAlumnesModificats=0
    info_nMissatgesEnviats = 0
    info_nResponsablesCreats = 0
    info_nRespSenseDni = 0
    info_nMenorsSenseResp = 0

    AlumnesCanviatsDeGrup = []
    AlumnesInsertats = []


    trobatGrupClasse = False
    trobatNom = False
    trobatDataNeixement = False
    trobatRalc = False

    # Carregar full de càlcul
    full = wb2.active
    max_row = full.max_row

    # iterar sobre les files
    colnames = [u'Identificador de l’alumne/a', u'Primer Cognom', u'Segon Cognom', u'Nom', u'Data naixement',
                u'Tutor 1 - 1r cognom', u'Tutor 1 - 2n cognom', u'Tutor 1 - nom', u'Contacte 1er tutor alumne - Valor',
                u'Tutor 2 - 1r cognom', u'Tutor 2 - 2n cognom', u'Tutor 2 - nom', u'Contacte 2on tutor alumne - Valor',
                u'Tipus de via', u'Nom via', u'Número', u'Bloc', u'Escala', u'Planta', u'Porta', u'Codi postal',
                u'Localitat de residència', u'Municipi de residència', u'Correu electrònic',
                u'Contacte altres alumne - Valor',
                u'Grup Classe', 'Número de document d’identitat', 'Tipus de document d’identitat', 
                'Tutor 1 - doc. identitat', 'Tutor 1 - tipus via', 'Tutor 1 - nom via', 'Tutor 1 - número', 
                'Tutor 1 - bloc', 'Tutor 1 - escala', 'Tutor 1 - planta', 'Tutor 1 - porta', 'Tutor 1 - CP', 
                'Tutor 1 - localitat', 'Tutor 1 - municipi', 
                'Tutor 2 - doc. identitat', 'Tutor 2 - tipus via', 'Tutor 2 - nom via', 'Tutor 2 - número', 
                'Tutor 2 - bloc', 'Tutor 2 - escala', 'Tutor 2 - planta', 'Tutor 2 - porta', 'Tutor 2 - CP', 
                'Tutor 2 - localitat', 'Tutor 2 - municipi',
                ]
    rows = list(wb2.active.rows)
    col_indexs = {n: cell.value.strip() for n, cell in enumerate(rows[5])
                   if cell.value and cell.value.strip() in colnames} # Començar a la fila 6, les anteriors són brossa
    cursos = set()
    for row in rows[6:max_row - 1]:  # la darrera fila també és brossa
        a = Alumne()
        a.ralc = ''
        a.altres_telefons = ''
        r1 = {}
        r2 = {}
        for index, cell in enumerate(row):
            if bool(cell) and bool(cell.value) and isinstance(cell.value, str):
                cell.value=cell.value.strip()
            if index in col_indexs:
                if col_indexs[index].endswith(u"Identificador de l’alumne/a"):
                    a.ralc=''.join(filter(str.isdigit, unicode(cell.value)))
                    if not bool(a.ralc): break # Salta línies sense dades
                    trobatRalc = True
                if col_indexs[index].endswith(u"Primer Cognom"):
                    a.cognoms = unicode(cell.value)
                if col_indexs[index].endswith(u"Segon Cognom"):
                    a.cognoms += (" " + unicode(cell.value)) if cell.value else ""
                if col_indexs[index].endswith(u"Nom"):
                    a.nom = unicode(cell.value)
                    trobatNom = True
                if col_indexs[index].endswith(u"Grup Classe"):
                    try:
                        unGrup = Grup2Aula.objects.get(grup_esfera = unicode(cell.value), Grup2Aula__isnull = False)
                        a.grup = unGrup.Grup2Aula
                    except:
                        return { 'errors': [ u"error carregant {0}".format( unicode(cell.value) ), ], 'warnings': [], 'infos': [] }
                    trobatGrupClasse = True
                if col_indexs[index].endswith(u"Correu electrònic"):
                    a.correu = unicode(cell.value) if cell.value else ""
                if col_indexs[index].endswith(u"Data naixement"):
                    try:
                        data = unicode(cell.value).split(" ")[0]
                        if "/" in data: dia = time.strptime(data, '%d/%m/%Y')
                        else: dia = time.strptime(data, '%Y-%m-%d')
                        a.data_neixement = datetime.fromtimestamp(time.mktime(dia)).date()
                        trobatDataNeixement = True
                    except Exception as e:
                        a.data_neixement = None
                        errors.append( "Data de naixement incorrecte '{0}' de l'alumne {1} {2} ({3}).".format(str(cell.value), 
                                                            a.nom, a.cognoms, a.ralc) )
                if col_indexs[index].endswith(u"Localitat de residència"):
                    a.localitat = unicode(cell.value) if cell.value else ""
                if col_indexs[index].endswith(u"Codi postal"):
                    a.cp = unicode(cell.value) if cell.value else ""
                if col_indexs[index].endswith(u"Municipi de residència"):
                    a.municipi = unicode(cell.value) if cell.value else ""
                if col_indexs[index].endswith(u"Contacte 1er tutor alumne - Valor"):
                    dades_tutor1 = dades_responsable(unicode(cell.value) if cell.value else "")
                    if dades_tutor1["mobils"]:
                        r1["telefon"] = ', '.join(dades_tutor1["mobils"])
                        a.altres_telefons = posarDada(dades_tutor1["fixes"], a.altres_telefons)
                    elif dades_tutor1["fixes"]:
                        r1["telefon"] = ', '.join(dades_tutor1["fixes"])
                    r1["correu"] = ', '.join(dades_tutor1["mails"])
                if col_indexs[index].endswith(u"Contacte 2on tutor alumne - Valor"):
                    dades_tutor2 = dades_responsable(unicode(cell.value) if cell.value else "")
                    if dades_tutor2["mobils"]:
                        r2["telefon"] = ', '.join(dades_tutor2["mobils"])
                        a.altres_telefons = posarDada(dades_tutor2["fixes"], a.altres_telefons)
                    elif dades_tutor2["fixes"]:
                        r2["telefon"] = ', '.join(dades_tutor2["fixes"])
                    r2["correu"] = ', '.join(dades_tutor2["mails"])
                if col_indexs[index].endswith(u"Contacte altres alumne - Valor"):
                    telefons_alumne = dades_responsable(unicode(cell.value) if cell.value else "")
                    a.altres_telefons = posarDada(telefons_alumne["fixes"], a.altres_telefons)
                    a.altres_telefons = posarDada(telefons_alumne["mobils"], a.altres_telefons)
                if col_indexs[index].endswith(u"Tutor 1 - 1r cognom"):
                    r1["cognoms"] = unicode(cell.value) if cell.value else ""
                if col_indexs[index].endswith(u"Tutor 1 - 2n cognom"):
                    r1["cognoms"] += (" " +  unicode(cell.value)) if cell.value else ""
                if col_indexs[index].endswith(u"Tutor 1 - nom"):
                    r1["nom"] = unicode(cell.value) if cell.value else ""
                if col_indexs[index].endswith(u"Tutor 2 - 1r cognom"):
                    r2["cognoms"] = unicode(cell.value) if cell.value else ""
                if col_indexs[index].endswith(u"Tutor 2 - 2n cognom"):
                    r2["cognoms"] += (" " +  unicode(cell.value)) if cell.value else ""
                if col_indexs[index].endswith(u"Tutor 2 - nom"):
                    r2["nom"] = unicode(cell.value) if cell.value else ""

                if col_indexs[index].endswith(u"Tipus de via"):
                    a.adreca = unicode(cell.value) if cell.value else ""
                if col_indexs[index].endswith(u"Nom via"):
                    a.adreca += " " +  unicode(cell.value) if cell.value else ""
                if col_indexs[index].endswith(u"Número"):
                    a.adreca += " " +   unicode(cell.value) if cell.value else ""
                if col_indexs[index].endswith(u"Bloc"):
                    a.adreca += " " +   unicode(cell.value) if cell.value else ""
                if col_indexs[index].endswith(u"Escala"):
                    a.adreca += " " +   unicode(cell.value) if cell.value else ""
                if col_indexs[index].endswith(u"Planta"):
                    a.adreca += " " +   unicode(cell.value) if cell.value else ""
                if col_indexs[index].endswith(u"Porta"):
                    a.adreca += " " +   unicode(cell.value) if cell.value else ""

                if col_indexs[index].endswith('Tutor 1 - doc. identitat'):
                    if cell.value and " - " in cell.value:
                        cell.value = cell.value.split(" - ")[0]
                    r1["dni"] = unicode(cell.value)[-10:] if cell.value else ""
                if col_indexs[index].endswith('Tutor 1 - tipus via'):
                    r1["adreca"] = (unicode(cell.value) if cell.value else "") + r1.get("adreca","")
                if col_indexs[index].endswith('Tutor 1 - nom via'):
                    r1["adreca"] = r1.get("adreca","") + (( " " + unicode(cell.value)) if cell.value else "")
                if col_indexs[index].endswith('Tutor 1 - número'):
                    r1["adreca"] = r1.get("adreca","") + (( " " + unicode(cell.value)) if cell.value else "")
                if col_indexs[index].endswith('Tutor 1 - bloc'):
                    r1["adreca"] = r1.get("adreca","") + (( " " + unicode(cell.value)) if cell.value else "")
                if col_indexs[index].endswith('Tutor 1 - escala'):
                    r1["adreca"] = r1.get("adreca","") + (( " " + unicode(cell.value)) if cell.value else "")
                if col_indexs[index].endswith('Tutor 1 - planta'):
                    r1["adreca"] = r1.get("adreca","") + (( " " + unicode(cell.value)) if cell.value else "")
                if col_indexs[index].endswith('Tutor 1 - porta'):
                    r1["adreca"] = r1.get("adreca","") + (( " " + unicode(cell.value)) if cell.value else "")
                if col_indexs[index].endswith('Tutor 1 - CP'):
                    r1["cp"] = unicode(cell.value) if cell.value else ""
                if col_indexs[index].endswith('Tutor 1 - localitat'):
                    r1["localitat"] = unicode(cell.value) if cell.value else ""
                if col_indexs[index].endswith('Tutor 1 - municipi'):
                    r1["municipi"] = unicode(cell.value) if cell.value else ""
                if col_indexs[index].endswith('Tutor 2 - doc. identitat'):
                    if cell.value and " - " in cell.value:
                        cell.value = cell.value.split(" - ")[0]
                    r2["dni"] = unicode(cell.value)[-10:] if cell.value else ""
                if col_indexs[index].endswith('Tutor 2 - tipus via'):
                    r2["adreca"] = (unicode(cell.value) if cell.value else "") + r2.get("adreca","")
                if col_indexs[index].endswith('Tutor 2 - nom via'):
                    r2["adreca"] = r2.get("adreca","") + (( " " + unicode(cell.value)) if cell.value else "")
                if col_indexs[index].endswith('Tutor 2 - número'):
                    r2["adreca"] = r2.get("adreca","") + (( " " + unicode(cell.value)) if cell.value else "")
                if col_indexs[index].endswith('Tutor 2 - bloc'):
                    r2["adreca"] = r2.get("adreca","") + (( " " + unicode(cell.value)) if cell.value else "")
                if col_indexs[index].endswith('Tutor 2 - escala'):
                    r2["adreca"] = r2.get("adreca","") + (( " " + unicode(cell.value)) if cell.value else "")
                if col_indexs[index].endswith('Tutor 2 - planta'):
                    r2["adreca"] = r2.get("adreca","") + (( " " + unicode(cell.value)) if cell.value else "")
                if col_indexs[index].endswith('Tutor 2 - porta'):
                    r2["adreca"] = r2.get("adreca","") + (( " " + unicode(cell.value)) if cell.value else "")
                if col_indexs[index].endswith('Tutor 2 - CP'):
                    r2["cp"] = unicode(cell.value) if cell.value else ""
                if col_indexs[index].endswith('Tutor 2 - localitat'):
                    r2["localitat"] = unicode(cell.value) if cell.value else ""
                if col_indexs[index].endswith('Tutor 2 - municipi'):
                    r2["municipi"] = unicode(cell.value) if cell.value else ""

        if not bool(a.ralc): continue # Salta línies sense dades
        info_nAlumnesLlegits += 1

        if not (trobatGrupClasse and trobatNom and trobatDataNeixement and trobatRalc):
            errors.append( u'Falten camps al fitxer' )
            return { 'errors': errors, 'warnings': [], 'infos': [] }

        alumneDadesAnteriors = None
        try:
            q_mateix_ralc = Q( ralc = a.ralc )
            alumneDadesAnteriors =Alumne.objects.get (q_mateix_ralc)

        except Alumne.DoesNotExist:
            pass

        if alumneDadesAnteriors is None:
            a.estat_sincronitzacio = 'S-I'
            a.data_alta = date.today()
            a.motiu_bloqueig = u'No sol·licitat'

            info_nAlumnesInsertats+=1
            AlumnesInsertats.append(a)

        else:
            #TODO: si canvien dades importants avisar al tutor.
            a.pk = alumneDadesAnteriors.pk
            a.nom_sentit = alumneDadesAnteriors.nom_sentit
            a.estat_sincronitzacio = 'S-U'
            info_nAlumnesModificats+=1

            # En cas que l'alumne pertanyi a un dels grups parametritzat com a estàtic,
            # no se li canviarà de grup en les importacions d'Esfer@.
            grups_estatics, _ = ParametreEsfera.objects.get_or_create( nom_parametre = 'grups estatics' )
            es_de_grup_estatic = False
            for prefixe_grup in grups_estatics.valor_parametre.split(','):
                prefix = prefixe_grup.replace(' ','')
                if prefix:
                    es_de_grup_estatic = es_de_grup_estatic or alumneDadesAnteriors.grup.descripcio_grup.startswith( prefix )

            if a.grup.pk != alumneDadesAnteriors.grup.pk:
                if es_de_grup_estatic: #no canviar-li de grup
                    a.grup = alumneDadesAnteriors.grup
                else:
                    AlumnesCanviatsDeGrup.append(a)

            a.user_associat = alumneDadesAnteriors.user_associat
            a.usuaris_app_associats.set(alumneDadesAnteriors.usuaris_app_associats.all())
            a.periodicitat_faltes = alumneDadesAnteriors.periodicitat_faltes
            a.periodicitat_incidencies = alumneDadesAnteriors.periodicitat_incidencies
            a.foto = alumneDadesAnteriors.foto
            a.responsable_preferent = alumneDadesAnteriors.responsable_preferent
            a.observacions = alumneDadesAnteriors.observacions
            #el recuperem, havia estat baixa:
            if alumneDadesAnteriors.data_baixa:
                info_nAlumnesInsertats+=1
                info_nAlumnesModificats-=1
                a.data_alta = date.today()
                a.motiu_bloqueig = ""
            else:
                a.data_alta = alumneDadesAnteriors.data_alta
                a.motiu_bloqueig = alumneDadesAnteriors.motiu_bloqueig
                
        #DEPRECATED vvv
        if alumneDadesAnteriors and alumneDadesAnteriors.correu_relacio_familia_pare:
            r1["correu_relacio_familia"]=alumneDadesAnteriors.correu_relacio_familia_pare
            r1["periodicitat_faltes"]=alumneDadesAnteriors.periodicitat_faltes
            r1["periodicitat_incidencies"]=alumneDadesAnteriors.periodicitat_incidencies
            a.correu_relacio_familia_pare = ''
        if alumneDadesAnteriors and alumneDadesAnteriors.correu_relacio_familia_mare:
            r2["correu_relacio_familia"]=alumneDadesAnteriors.correu_relacio_familia_mare
            r2["periodicitat_faltes"]=alumneDadesAnteriors.periodicitat_faltes
            r2["periodicitat_incidencies"]=alumneDadesAnteriors.periodicitat_incidencies
            a.correu_relacio_familia_mare = ''
        #DEPRECATED ^^^

        set_aruco_marker(a, markers_per_nivell[a.grup.curs.nivell.pk])

        a.save()
        # Crea usuaris Responsable
        err_resp_menors = 0
        if (r1.get("cognoms",None) or r1.get("nom",None)) and not r1.get("dni",None):
            info_nRespSenseDni += 1
            if a.edat()<18: err_resp_menors += 1
        if (r2.get("cognoms",None) or r2.get("nom",None)) and not r2.get("dni",None):
            info_nRespSenseDni += 1
            if a.edat()<18: err_resp_menors += 1
        if err_resp_menors==2: info_nMenorsSenseResp += 1
        info_nResponsablesCreats += creaResponsables(a, [r1, r2])
        cursos.add(a.grup.curs)
        #DEPRECATED vvv
        if alumneDadesAnteriors and not alumneDadesAnteriors.responsable_preferent and alumneDadesAnteriors.responsables.exists():
            resp1, resp2 = a.get_responsables()
            if alumneDadesAnteriors.primer_responsable==1 and resp2: a.responsable_preferent = resp2
            else: a.responsable_preferent = resp1
            a.save()
        #DEPRECATED ^^^
        
    #
    # Baixes:
    #
    # Els alumnes de Saga no s'han de tenir en compte per fer les baixes
    AlumnesDeSaga = Alumne.objects.exclude(grup__curs__in=cursos)
    # Es canvia estat PRC a ''. No modifica DEL ni MAN
    AlumnesDeSaga.filter( estat_sincronitzacio__exact = 'PRC' ).update(estat_sincronitzacio='')
    
    #Els alumnes que hagin quedat a PRC és que s'han donat de baixa:
    AlumnesDonatsDeBaixa = Alumne.objects.filter( estat_sincronitzacio__exact = 'PRC' )
    AlumnesDonatsDeBaixa.update(
                            data_baixa = date.today(),
                            estat_sincronitzacio = 'DEL' ,
                            motiu_bloqueig = 'Baixa'
                            )

    #Avisar als professors: Baixes
    #: enviar un missatge a tots els professors que tenen aquell alumne.
    info_nAlumnesEsborrats = len(  AlumnesDonatsDeBaixa )
    professorsNotificar = {}
    for alumneDonatDeBaixa in AlumnesDonatsDeBaixa:
        for professor in Professor.objects.filter(  horari__impartir__controlassistencia__alumne = alumneDonatDeBaixa ).distinct():
            professorsNotificar.setdefault( professor.pk, []  ).append( alumneDonatDeBaixa )
    for professorPK, alumnes in professorsNotificar.items():
        llista = []
        for alumne in alumnes:
            llista.append( u'{0} ({1})'.format(unicode( alumne), alumne.grup.descripcio_grup ) )
        missatge = ALUMNES_DONATS_DE_BAIXA
        tipus_de_missatge = tipusMissatge(missatge)
        msg = Missatge( remitent = user, text_missatge = missatge, tipus_de_missatge = tipus_de_missatge  )
        msg.afegeix_infos( llista )
        msg.envia_a_usuari( Professor.objects.get( pk = professorPK ) , 'IN')
        info_nMissatgesEnviats += 1

    #Avisar als professors: Canvi de grup
    #enviar un missatge a tots els professors que tenen aquell alumne.
    info_nAlumnesCanviasDeGrup = len(  AlumnesCanviatsDeGrup )
    professorsNotificar = {}
    for alumneCanviatDeGrup in AlumnesCanviatsDeGrup:
        qElTenenALHorari = Q( horari__impartir__controlassistencia__alumne = alumneCanviatDeGrup   )
        qImparteixDocenciaAlNouGrup = Q(  horari__grup =  alumneCanviatDeGrup.grup )
        for professor in Professor.objects.filter( qElTenenALHorari | qImparteixDocenciaAlNouGrup  ).distinct():
            professorsNotificar.setdefault( professor.pk, []  ).append( alumneCanviatDeGrup )
    for professorPK, alumnes in professorsNotificar.items():
        llista = []
        for alumne in alumnes:
            llista.append( u'{0} passa a grup {1}'.format(unicode( alumne), alumne.grup.descripcio_grup ) )
        missatge = ALUMNES_CANVIATS_DE_GRUP
        tipus_de_missatge = tipusMissatge(missatge)
        msg = Missatge( remitent = user, text_missatge = missatge, tipus_de_missatge = tipus_de_missatge  )
        msg.afegeix_infos( llista )
        msg.envia_a_usuari( Professor.objects.get( pk = professorPK ) , 'IN')
        info_nMissatgesEnviats += 1

    #Avisar als professors: Altes
    #enviar un missatge a tots els professors que tenen aquell alumne.
    professorsNotificar = {}
    for alumneNou in AlumnesInsertats:
        qImparteixDocenciaAlNouGrup = Q(  horari__grup =  alumneNou.grup )
        for professor in Professor.objects.filter( qImparteixDocenciaAlNouGrup ).distinct():
            professorsNotificar.setdefault( professor.pk, []  ).append( alumneNou )
    for professorPK, alumnes in professorsNotificar.items():
        llista = []
        for alumne in alumnes:
            llista.append( u'{0} al grup {1}'.format(unicode( alumne), alumne.grup.descripcio_grup ) )
        missatge = ALUMNES_DONATS_DALTA
        tipus_de_missatge = tipusMissatge(missatge)
        msg = Missatge( remitent = user, text_missatge = missatge, tipus_de_missatge = tipus_de_missatge  )
        msg.afegeix_infos( llista )
        msg.envia_a_usuari( Professor.objects.get( pk = professorPK ) , 'IN')
        info_nMissatgesEnviats += 1


    #Treure'ls de les classes: les baixes
    ControlAssistencia.objects.filter(
                impartir__dia_impartir__gte = date.today(),
                alumne__in = AlumnesDonatsDeBaixa ).delete()

    #Treure'ls de les classes: els canvis de grup   #Todo: només si l'àmbit és grup.

    ambit_no_es_el_grup = Q( impartir__horari__assignatura__tipus_assignatura__ambit_on_prendre_alumnes__in = [ 'C', 'N', 'I' ] )
    ( ControlAssistencia
      .objects
      .filter( ambit_no_es_el_grup )
      .filter( impartir__dia_impartir__gte = date.today() )
      .filter( alumne__in = AlumnesCanviatsDeGrup )
      .delete()
     )


    #Altes: posar-ho als controls d'assistència de les classes (?????????)


    #
    # FI
    #
    # Tornem a l'estat de sincronització en blanc, excepte els alumnes esborrats DEL i els alumnes entrats manualment MAN.
    Alumne.objects.exclude( estat_sincronitzacio__exact = 'DEL' ).exclude( estat_sincronitzacio__exact = 'MAN') \
        .update( estat_sincronitzacio = '')
    errors.append( u'%d alumnes sense grup'%errors_nAlumnesSenseGrup )
    warnings= [  ]
    infos=    [   ]
    infos.append(u'{0} alumnes llegits'.format(info_nAlumnesLlegits) )
    infos.append(u'{0} alumnes insertats'.format(info_nAlumnesInsertats) )
    infos.append(u'{0} alumnes modificats'.format(info_nAlumnesModificats ) )
    infos.append(u'{0} alumnes esborrats'.format(info_nAlumnesEsborrats ) )
    infos.append(u'{0} alumnes canviats de grup'.format(info_nAlumnesCanviasDeGrup ) )
    infos.append(u'{0} alumnes en estat sincronització manual'.format( \
        len(Alumne.objects.filter(estat_sincronitzacio__exact = 'MAN'))))
    infos.append(u'{0} missatges enviats'.format(info_nMissatgesEnviats ) )
    infos.append(u'{0} responsables creats'.format(info_nResponsablesCreats ) )
    if info_nRespSenseDni>0: infos.append(u'{0} responsables sense identificació'.format(info_nRespSenseDni ) )
    if info_nMenorsSenseResp>0: infos.append(u'{0} alumnes menors d\'edat sense responsable'.format(info_nMenorsSenseResp ) )
    missatge = IMPORTACIO_ESFERA_FINALITZADA
    tipus_de_missatge = tipusMissatge(missatge)
    msg = Missatge(
                remitent= user,
                text_missatge = missatge,
                tipus_de_missatge = tipus_de_missatge)
    msg.afegeix_errors( errors )
    msg.afegeix_warnings(warnings)
    msg.afegeix_infos(infos)
    importancia = 'VI' if len( errors )> 0 else 'IN'
    grupDireccio =  Group.objects.get( name = 'direcció' )
    msg.envia_a_grup( grupDireccio , importancia=importancia)

    return { 'errors': errors, 'warnings': warnings, 'infos': infos }




def comprovar_grups( f ):
    errors = []
    warnings = []
    infos = []

    # Carregar full de càlcul
    wb2 = load_workbook(f)
    full = wb2.active
    max_row = full.max_row
    max_column = full.max_column

    # iterar sobre les files
    colname = [u'Grup Classe']
    rows = list(wb2.active.rows)
    col_index = {n: cell.value for n, cell in enumerate(rows[5])
                   if cell.value in colname}  # Començar a la fila 6, les anteriors són brossa

    if col_index is None:
        errors.append(u"No trobat el grup classe al fitxer d'importació")
        return { 'errors': errors, 'warnings': warnings, 'infos': infos }

    for row in rows[6:max_row - 1]:  # la darrera fila també és brossa
        for index, cell in enumerate(row):
            if index in col_index:
                grup_classe = unicode(cell.value).strip()
                _, new = Grup2Aula.objects.get_or_create(grup_esfera=grup_classe)
                if new:
                    errors.append(
                        u"El grup '{grup_classe}' de l'Esfer@ no té correspondència al programa. Revisa les correspondències Esfer@-Aula".format(
                            grup_classe=grup_classe))

    return { 'errors': errors }



def dades_responsable ( dades ):
    splitted = dades.split(" - ")
    mails = [ dada for dada in splitted if "@" in dada ]
    fixes = [ dada for dada in splitted if dada.startswith(("9","8")) and dada not in mails ]
    mobils =[ dada for dada in splitted if dada not in mails+fixes ]

    dades_tutor = { "mails": mails,
                   "fixes": fixes,
                   "mobils": mobils,
                 }
    return dades_tutor


def dades_adiccionals (f, user=None):
    errors = []
    warnings = []
    camps_addicionals = [x['label'] for x in CUSTOM_DADES_ADDICIONALS_ALUMNE]

    try:
        # Carregar full de càlcul
        wb2 = load_workbook(f)
        if len(wb2.worksheets) != 1:
            # Si té més d'una pestanya --> error
            errors.append('Fitxer incorrecte')
            return {'errors': errors, 'warnings': [], 'infos': []}
    except:
        errors.append('Fitxer incorrecte')
        return {'errors': errors, 'warnings': [], 'infos': []}

    info_nLiniesLlegides = 0
    info_nModificacions = 0

    # Carregar full de càlcul
    full = wb2.active
    max_row = full.max_row

    # iterar sobre les files
    colnames = [u'Identificador de l’alumne/a', u'Nom complet de l’alumne/a', u'Camps lliures - Nom',
                u'Camps lliures - Valor']
    rows = list(wb2.active.rows)
    col_indexs = {n: cell.value for n, cell in enumerate(rows[9])
                  if cell.value in colnames}  # Començar a la fila 10, les anteriors són brossa
    for row in rows[10:max_row - 1]:  # la darrera fila també és brossa
        info_nLiniesLlegides += 1
        ralc_llegit = ''
        nom_llegit = ''
        alumne = None
        for index, cell in enumerate(row):
            if bool(cell) and bool(cell.value) and isinstance(cell.value, str):
                cell.value = cell.value.strip()
            if index in col_indexs:
                if col_indexs[index].endswith(u"Camps lliures - Nom"):
                    camp = unicode(cell.value)
                    if camp in camps_addicionals:
                        ralc_llegit = unicode(cell.offset(0,-2).value) #accedim a la columna on es troba el ralc
                        try:
                            alumne = Alumne.objects.get(ralc=ralc_llegit)
                        except:
                            return {'errors': [u"error carregant, Ralc {0} no trobat - línia {1} del fitxer de càrrega".format(ralc_llegit,info_nLiniesLlegides+10), ],
                                    'warnings': [],'infos': []}
                        nom_llegit = unicode(cell.offset(0,-1).value) #accedim a la columna on es troba el nom
                        if alumne.__str__() != nom_llegit:
                            warnings.append(u'Nom/Cognoms alumne/a amb ralc {0} no coincident: {1} - {2}'.format(alumne.ralc, alumne.__str__(), nom_llegit))
                        try:
                            valor = unicode(cell.offset(0,1).value) if cell.offset(0,1).value else ""  #accedim a la següent columna on està el valor corresponent al camp
                            dada_addicional, created = DadesAddicionalsAlumne.objects.get_or_create(alumne=alumne, label=camp)
                            if dada_addicional.value != valor:
                                dada_addicional.value = valor
                                dada_addicional.save()
                                info_nModificacions += 1
                        except:
                            return {'errors': [
                                u"error carregant, línia {0} del fitxer de càrrega".format(info_nLiniesLlegides+10), ],
                                    'warnings': [],
                                    'infos': []}

    infos = ['Resum:']
    infos.append(u'{0} línies llegides'.format(info_nLiniesLlegides))
    infos.append(u'{0} modificacions realitzades'.format(info_nModificacions))
    if warnings: warnings.insert(0, "Avisos:")
    missatge = IMPORTACIO_DADES_ADDICIONALS_FINALITZADA
    tipus_de_missatge = tipusMissatge(missatge)
    msg = Missatge(
        remitent=user,
        text_missatge=missatge,
        tipus_de_missatge=tipus_de_missatge)
    msg.afegeix_errors(errors)
    msg.afegeix_warnings(warnings)
    msg.afegeix_infos(infos)
    importancia = 'VI' if len(errors) > 0 else 'IN'
    grupDireccio = Group.objects.get(name='direcció')
    msg.envia_a_grup(grupDireccio, importancia=importancia)

    return {'errors': errors, 'warnings': warnings, 'infos': infos}

